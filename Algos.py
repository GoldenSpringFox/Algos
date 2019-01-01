"""
This is the code to the program Algos.
Algos is a recipe adviser program. for a given spectrum, it recommends changes in the tool parameters in order to
achieve target spectrum.
All copyrights go to Rioglass
Programmed by: Aviv Goldstein
Developed by: Alona Goldstein
"""

# creating exe instructions:
# open cmd, navigate to the codes folder using ->cd <PATH>
# tkinter command: pyinstaller.exe --onefile --windowed --icon=Algos_icon_1.ico Algos.py
# in the folder dist there will be the exe
# to mail it, use zip

# ~~~~~~~~~~~~~~~~~~~~~ Imports ~~~~~~~~~~~~~~~~~~~~~~~~

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from operator import itemgetter
from decimal import Decimal
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import matplotlib as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
import pickle
from pyperclip import copy as pypcopy
from ordered_set import OrderedSet
# import os

plt.use("TkAgg")

# ~~~~~~~~~~~~~~~~~ Global variables ~~~~~~~~~~~~~~~~~

# Constant variables (even though python doesn't have these)
MIN_FACTOR = 2
MAX_FACTOR = 23
MIN_WAVELENGTH = 400
MAX_WAVELENGTH = 2000
STEP_WAVELENGTH = 10
SCREEN_WIDTH = None
SCREEN_HEIGHT = None
MAX_NUM_OF_LOOPS = 10
FACTORS = ["I₁", "I₂", "I₃", "I₄", "I₅", "I₆", "V₁₋₂", "V₃₋₄", "V₅₋₆", "Vₐᵣ"]
COLOR_TITLE = "#009999"
COLOR_BACKGROUND = "#f2f2f2"
COLOR_SECONDARY = "#0e403f"
COLOR_BOTTOM = "#a6a6a6"
ALGOS_PATH = r"C:\Program Files\AlgosApp"
HELP_TEXT = "ALGOS is a recipe adviser tool. It was designed to help the process engineer to tune the coating " \
            "parameters of a given run, to reach a target optical spectrum." \
            "\nALGOS works with the aid of an excel file, which contains the following datasheets:" \
            "\n\t1.\tData" \
            "\n\t2.\tTarget" \
            "\n\t3.\tFactors" \
            "\n\t4.\tSolarSpectrum" \
            "\n\t5.\tResults" \
            "\nStage1: paste the measured reflectance data of the run that you wish to optimize in the “Data” " \
            "worksheet. Save and close the file." \
            "\nStage2: open ALGOS and choose the excel file that you wish to work with." \
            "\nStage3: choose between two modes of operation:" \
            "\nMode I - manual change of specific parameters to simulate their influence on the spectrum; " \
            "\n         observe the resultant graphs and estimated change in alpha and epsilon on a separate graph " \
            "window." \
            "\nMode II – calculation of the recommended process changes that will bring the spectrum closest to " \
            "the target." \
            "\n          The calculated spectra will appear in “Results” datasheet and can be also copied by pressing" \
            " “copy data” from the graph window" \
            "\n         (please note that the copied data is of R between 400-1990 nm, steps 10 nm)." \
            "\n\nEnjoy!"
ABOUT_TEXT = "ALGOS version 1.0" \
             "\nProgrammed by: Aviv Goldstein" \
             "\nModelled by: Dr. Alona Goldstein" \
             "\n© Rioglass Solar Systems Ltd." \
             "\nDec 2018"

# Global Variables (Remove this)
global_best_results = []


# ~~~~~~~~~~~~~~~~~~~~~ Classes ~~~~~~~~~~~~~~~~~~~~~

###################################################################################
class GraphFrame(tk.Toplevel):

    # ------------------------------------------------------
    def __init__(self, original, spectrum, factors, dev, alpha, epsilon):
        """Constructor"""
        self.original_frame = original
        self.spectrum = spectrum
        self.factors = factors
        self.deviation = dev
        self.d_alpha = calculate_difference(alpha, self.original_frame.original_alpha)
        self.d_epsilon = calculate_difference(epsilon, self.original_frame.original_epsilon, 1)
        tk.Toplevel.__init__(self)
        self.title(factors)

        # rest of the configuration code. set the layout of the graph screen
        f = Figure(figsize=(5, 5), dpi=100)
        a = f.add_subplot(111)
        a.plot(list(self.original_frame.target_lower_limit.keys()),
               list(self.original_frame.target_lower_limit.values()), '--', color="red")
        a.plot(list(self.original_frame.target_upper_limit.keys()),
               list(self.original_frame.target_upper_limit.values()), '--', color="red", label='_nolegend_')
        a.plot(list(self.original_frame.my_dict.keys()), list(self.original_frame.my_dict.values()), color="blue")
        if factors != "Original Spectrum":
            a.plot(list(spectrum.keys()), list(spectrum.values()), color="lime")
        a.set_xlabel("Wavelength[nm]")
        a.set_ylabel("Reflectance[%]")
        a.legend(["target limits", "original spectrum", "simulated spectrum"])

        self.top_frame = tk.Frame(self, bg=COLOR_TITLE)
        self.top_frame.grid(sticky="nsew")
        label_text = "Factors used: " + self.factors + "\nDeviation: %.4f" % self.deviation + \
                     "\nAlpha change: " + self.d_alpha + "\nEpsilon change: " + self.d_epsilon
        label = tk.Label(self.top_frame, text=label_text, anchor=tk.W, justify=tk.LEFT, font="Calibri 14 bold",
                         bg=COLOR_TITLE)
        label.pack(side=tk.LEFT)

        self.spectrum_text = ""
        for i in self.spectrum:
            self.spectrum_text += str(self.spectrum[i]) + "\n"
        button_copy_spectrum_to_clipboard = tk.Button(self.top_frame, text="copy data",
                                                      command=pypcopy(self.spectrum_text))
        button_copy_spectrum_to_clipboard.pack(side=tk.RIGHT, padx=(0, 40))

        canvas_frame = tk.Frame(self)
        canvas_frame.grid(row=1, columnspan=2, sticky="nsew")
        canvas = FigureCanvasTkAgg(f, canvas_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        toolbar = NavigationToolbar2Tk(canvas, canvas_frame)
        toolbar.config(bg=COLOR_SECONDARY)
        toolbar._message_label.config(background=COLOR_SECONDARY)
        toolbar.update()
        canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)


###################################################################################
class TextFrame(tk.Toplevel):
    def __init__(self, original, title, text):
        tk.Toplevel.__init__(self)
        self.original = original

        self.title(title)
        self.text = tk.Label(self, text=text, anchor=tk.NW, justify=tk.LEFT, font="Calibri 12", padx=10, pady=10)
        self.text.grid(sticky="nsew")
        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)

        center(self)


###################################################################################
class AlgosApp(object):

    # ----------------------------------------------------------------
    # Constructor
    def __init__(self, parent):
        # <><><><><><><> variable declaration <><><><><><><>
        self.best_results = [("", float("inf"), float("inf"))] * 5
        self.best_target_results = [("", float("inf"), float("inf"))] * 5
        self.my_wb = None
        self.data_ws = None
        self.factors_ws = None
        self.target_ws = None
        self.results_ws = None
        self.solar_spectrum_ws = None

        self.my_dict = None
        self.my_target = None
        self.my_solar_spectrum = None
        self.factor_places = None
        self.target_lower_limit = None
        self.target_upper_limit = None
        self.best_results_table = None
        self.original_alpha = None
        self.original_epsilon = None
        self.factors_set = OrderedSet([])

        self.graphs = []
        self.recommendation_chart = []

        # os.makedirs(os.path.dirname(ALGOS_PATH), exist_ok=True)

        self.new_file = 'mypickle.pk'  # i removed the ALGOS_PATH that was before the file name

        # <><><><><><><> Main window setup <><><><><><><>
        #########################
        # root and frames setup #
        #########################
        self.root = parent
        self.root.title("ALGOS")
        self.root.resizable(0, 0)
        # self.root.geometry("900x630")
        # self.root.config(bg=COLOR_SECONDARY)
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)  # runs on_close when the application closes

        self.window = tk.Frame(root, bg=COLOR_SECONDARY)
        self.window.pack_propagate(0)
        self.window.pack(fill=tk.BOTH, expand=1)
        self.window.columnconfigure(0, weight=1)
        self.window.rowconfigure(1, weight=1)
        self.window.bind("<Configure>", self.configure)

        self.menu = tk.Menu(self.window)
        self.root.config(menu=self.menu)

        self.top_frame = tk.Frame(self.window, bg=COLOR_TITLE)
        self.top_frame.grid(columnspan=3, sticky="nsew")

        self.main_frame = tk.Frame(self.window)
        self.main_frame.grid(row=1, column=1)
        self.main_frame.grid_remove()

        self.left_filler = tk.Frame(self.window, width=120, bg=COLOR_SECONDARY)
        self.left_filler.grid(row=1, sticky="nsew")

        self.factors_frame = tk.Frame(self.main_frame, padx=10,
                                      highlightbackground=COLOR_SECONDARY, highlightthickness=10)
        self.factors_frame.grid(row=0, column=1, sticky="nsew")

        self.middle_filler = tk.Frame(self.main_frame, width=50, bg=COLOR_SECONDARY)
        self.middle_filler.grid(row=0, column=2, sticky="nsew")

        self.recommendation_frame = tk.Frame(self.main_frame, padx=50,
                                             highlightbackground=COLOR_SECONDARY, highlightthickness=10)
        self.recommendation_frame.grid(row=0, column=3, sticky="nsew")

        self.right_filler = tk.Frame(self.window, width=120, bg=COLOR_SECONDARY)
        self.right_filler.grid(row=1, column=2, sticky="nsew")

        self.replacement_frame = tk.Frame(self.window, height=433, bg=COLOR_SECONDARY)
        self.replacement_frame.grid(row=1, column=1, sticky="nsew")

        self.bottom_frame = tk.Frame(self.window, bg=COLOR_BOTTOM)
        self.bottom_frame.grid(columnspan=3, sticky="nsew")

        ##############
        # menu stuff #
        ##############
        self.filemenu = tk.Menu(self.menu)
        self.menu.add_cascade(label="File", menu=self.filemenu)
        # self.filemenu.add_command(label="New")  # , command='function name'
        # self.filemenu.add_command(label="Open...")
        # self.filemenu.add_command(label="Optimize target", command=self.optimize_target)
        self.filemenu.add_command(label="About", command=lambda: self.open_text_window("About", ABOUT_TEXT))
        self.filemenu.add_separator()
        self.filemenu.add_command(label="Exit", command=root.destroy)

        self.helpmenu = tk.Menu(self.menu)
        self.menu.add_cascade(label="Help", menu=self.helpmenu)
        self.helpmenu.add_command(
            label="Help",
            command=lambda: self.open_text_window("Help", HELP_TEXT))

        ###########################
        # replacement_frame stuff #
        ###########################
        self.replacement_frame_temp = tk.Frame(self.replacement_frame, bg=COLOR_BACKGROUND, pady=100, padx=15)
        self.replacement_frame_temp.pack()
        self.replacement_label = tk.Label(self.replacement_frame_temp, bg=COLOR_BACKGROUND,
                                          text="In order to start, \nPlease choose the file you wish to work with",
                                          font="Calibri 20", padx=25, pady=50, anchor=tk.CENTER)
        self.file_select_button_temp = tk.Button(self.replacement_frame_temp, text="Choose file", font="Calibri 16",
                                                 command=self.open_file_explorer_window_button, pady=25, padx=25,
                                                 bg=COLOR_BACKGROUND)

        self.replacement_label.grid()
        self.file_select_button_temp.grid(row=1)

        ###################
        # top_frame stuff #
        ###################
        self.file_label = tk.Label(self.top_frame, text="Current File: ", font="Calibri 14 bold", anchor=tk.W)
        self.label_file_path = tk.StringVar()
        # self.label_file_path.set("None")
        self.main_label_text = tk.StringVar()
        # self.main_label_text.set("Welcome to ALGOS")
        try:
            with open(self.new_file, 'rb') as fi:
                text = pickle.load(fi)
                self.label_file_path.set(text)
        except FileNotFoundError:
            with open(self.new_file, 'wb') as fi:
                # dump your data into the file
                pickle.dump("", fi)
                self.label_file_path.set("")
        # self.open_file_explorer_window_button(False)
        self.file_select_label = tk.Label(
            self.top_frame, textvariable=self.label_file_path, font="Calibri 14",
            anchor=tk.W, borderwidth=2, relief="solid")  # , width=70
        self.file_select_button = tk.Button(
            self.top_frame, text="Choose file", command=self.open_file_explorer_window_button, pady=2)
        self.main_label = tk.Label(
            self.top_frame, compound=tk.CENTER, font="Calibri 18 bold", pady=20, padx=10, bg=COLOR_TITLE,
            textvariable=self.main_label_text)

        self.file_label.grid()
        self.file_select_label.grid(row=0, column=1, sticky="nsew")
        self.file_select_button.grid(row=0, column=2)
        self.main_label.grid(row=1, columnspan=3)

        # self.top_frame.grid_rowconfigure(2, weight=1)
        self.top_frame.grid_columnconfigure(1, weight=1)

        #######################
        # factors_frame stuff #
        #######################
        self.factors_frame_label = tk.Label(self.factors_frame, text="Manual Simulation", font="Calibri 16")
        self.factors_frame_label.grid()
        self.text_factors = tk.Text(self.factors_frame, height=3, width=30, bg="#e6ffff", font="Calibri 16")  # #00ffff
        self.text_factors.grid(row=1)

        self.container_frame1 = tk.Frame(self.factors_frame)
        self.container_frame1.grid(row=2)
        self.labels = []
        self.buttons = []
        self.initialize_buttons()

        self.container_frame2 = tk.Frame(self.factors_frame)
        self.container_frame2.grid(row=3, pady=(0, 2))
        self.original_graph_button = tk.Button(self.container_frame2, text="Original\nGraph", padx=20, width=2,
                                               command=lambda: self.open_graph_window(
                                                   self.my_dict, "Original Spectrum",
                                                   dev=compare_to_target(self.my_target, self.my_dict),
                                                   alpha=calculate_alpha(self.my_solar_spectrum, self.my_dict),
                                                   epsilon=calculate_epsilon(self.my_dict)))
        self.delete_factor_button = tk.Button(self.container_frame2, text="Delete", padx=20, pady=8, width=2,
                                              state="disabled", command=self.delete_factor_button_function)
        self.confirm_factor_button = tk.Button(self.container_frame2, text="Simulate", padx=40, pady=5, width=3,
                                               state="disabled", command=self.confirm_factor_button_function,
                                               font="Calibri 12 bold")

        self.delete_factor_button.grid()
        self.confirm_factor_button.grid(row=0, column=1)
        self.original_graph_button.grid(row=0, column=2)

        ##############################
        # recommendation_frame stuff #
        ##############################
        self.recommendation_frame_label = tk.Label(self.recommendation_frame, text="Recipe Adviser", font="Calibri 16")

        self.container_number_of_loops = tk.Frame(self.recommendation_frame)
        self.label_number_of_loops = tk.Label(self.container_number_of_loops, text="Number of factors: ",
                                              font="Calibri 12")
        self.number_of_loops = tk.Scale(self.container_number_of_loops, from_=1, to=6, orient=tk.HORIZONTAL)
        self.number_of_loops.set(3)

        self.button = tk.Button(
            self.recommendation_frame, text="Calculate", padx=40, pady=5, width=3, font="Calibri 12 bold",
            command=lambda: self.main_function(self.recommendation_frame_container))

        self.table_label = tk.Label(self.recommendation_frame, text="Recommended actions", font="Calibri 14")
        self.recommendation_frame_container = tk.Frame(self.recommendation_frame)

        # grid
        self.recommendation_frame_label.grid()

        self.container_number_of_loops.grid(row=1)
        self.label_number_of_loops.grid(pady=(15, 0))
        self.number_of_loops.grid(row=0, column=1)

        self.button.grid(row=2, padx=100, pady=(15, 25))
        self.table_label.grid(row=3)
        self.table_label.grid_remove()
        self.recommendation_frame_container.grid(row=4, pady=15, sticky="nsw")

        ######################
        # bottom_frame stuff #
        ######################
        self.bottom_label1 = tk.Label(self.bottom_frame, text="© Rioglass Solar Systems Ltd.",
                                      font="Calibri 10 bold", anchor=tk.W, bg=COLOR_BOTTOM)
        self.bottom_label1.grid()
        self.bottom_label2 = tk.Label(self.bottom_frame, text="version 1.0",
                                      font="Calibri 10 bold", bg=COLOR_BOTTOM)
        self.bottom_label2.grid(row=0, column=1)
        self.bottom_label3 = tk.Label(self.bottom_frame, text="Programmed by: Aviv Goldstein",
                                      font="Calibri 10 bold", anchor=tk.E, bg=COLOR_BOTTOM)
        self.bottom_label3.grid(row=0, column=2)

        self.bottom_frame.grid_columnconfigure(1, weight=1)

        # other stuff
        self.open_file_explorer_window_button(False)
        center(self.root)

    # ----------------------------------------------------------------
    def hide(self):
        """"""
        self.root.withdraw()

    # ----------------------------------------------------------------
    def show(self):
        """"""
        self.root.update()
        self.root.deiconify()

    # ----------------------------------------------------------------
    def open_graph_window(self, spectrum, factors, dev, alpha, epsilon):
        """"""
        self.graphs.append(GraphFrame(self, spectrum, factors, dev, alpha, epsilon))

    def open_text_window(self, title, text):
        TextFrame(self, title, text)

    # ----------------------------------------------------------------
    def open_file_explorer_window_button(self, is_button=True):
        """"""
        if is_button:
            self.label_file_path.set(filedialog.askopenfilename(initialdir="/", title="Select excel file",
                                                                filetypes=(("excel files", "*.xlsx"),)))
        try:
            self.open_file_worksheets()
        except KeyError:
            if is_button:
                messagebox.showwarning("Unsuitable file", "The file you chose does not match the required form.\n"
                                                          "Please make sure the file contains the Data, Factors, Target"
                                                          " and Results sheets")
                self.replacement_frame.grid()
                self.main_frame.grid_remove()
        except FileNotFoundError:
            self.label_file_path.set("None")
            self.main_label_text.set("Welcome to ALGOS")

            self.replacement_frame.grid()
            self.main_frame.grid_remove()
            self.left_filler.config(width=120)
            self.right_filler.config(width=120)

            for graph in self.graphs:
                graph.destroy()
        else:
            self.main_label_text.set("ALGOS")
            self.initialize_buttons()

            self.replacement_frame.grid_remove()
            self.main_frame.grid()
            self.left_filler.config(width=40)
            self.right_filler.config(width=40)

            for graph in self.graphs:
                graph.destroy()

            self.open_graph_window(self.my_dict, "Original Spectrum",
                                   dev=compare_to_target(self.my_target, self.my_dict),
                                   alpha=self.original_alpha,
                                   epsilon=self.original_epsilon)

    # ----------------------------------------------------------------
    # opens the requested excel file and saves all of the required sheets into class variables
    def open_file_worksheets(self):
        self.my_wb = load_workbook(self.label_file_path.get(), data_only=True)
        self.data_ws = self.my_wb["Data"]
        self.factors_ws = self.my_wb["Factors"]
        self.target_ws = self.my_wb["Target"]
        self.results_ws = self.my_wb["Results"]
        self.solar_spectrum_ws = self.my_wb["SolarSpectrum"]

        self.my_dict = data_reduce(self.data_ws)
        self.my_target = data_reduce(self.target_ws)
        self.my_solar_spectrum = data_reduce(self.solar_spectrum_ws)
        self.factor_places = map_factor_places(self.factors_ws)
        self.target_lower_limit = data_reduce(self.target_ws, 5)
        self.target_upper_limit = data_reduce(self.target_ws, 6)
        self.original_alpha = calculate_alpha(self.my_solar_spectrum, self.my_dict)
        self.original_epsilon = calculate_epsilon(self.my_dict)
        self.factors_set.clear()
        self.factors_set = find_factors(self.factors_ws, self.factors_set)

    def initialize_buttons(self):
        val = 0
        self.buttons.clear()
        self.container_frame1.destroy()
        self.container_frame1 = tk.Frame(self.factors_frame)
        self.container_frame1.grid(row=2)

        for val, factor in enumerate(self.factors_set):
            no_text = 1
            text = ""
            if factor[0] == "I":
                text = "1"
            elif factor[0] == "V":
                text = "0.5"
            else:
                no_text = 2
            self.buttons.append(tk.Button(
                self.container_frame1, text="-%g" % (float(text)*2) * no_text, padx=10, width=2,
                command=lambda c=val: self.factor_button_clicked(self.factors_set[c] + "--")))
            self.buttons[-1].grid(row=2 + val, column=0)
            self.buttons.append(tk.Button(self.container_frame1, text="-"+text, padx=10, width=2,
                                          command=lambda c=val: self.factor_button_clicked(self.factors_set[c] + "-")))
            self.buttons[-1].grid(row=2 + val, column=1)
            self.labels.append(tk.Label(self.container_frame1, text=factor, padx=25, font="Calibri 16"))
            self.labels[-1].grid(row=2 + val, column=2)
            self.buttons.append(tk.Button(self.container_frame1, text="+"+text, padx=10, width=2,
                                          command=lambda c=val: self.factor_button_clicked(self.factors_set[c] + "+")))
            self.buttons[-1].grid(row=2 + val, column=3)
            self.buttons.append(tk.Button(
                self.container_frame1, text="+%g" % (float(text)*2) * no_text, padx=10, width=2,
                command=lambda c=val: self.factor_button_clicked(self.factors_set[c] + "++")))
            self.buttons[-1].grid(row=2 + val, column=4)

    # ----------------------------------------------------------------
    # main function
    def main_function(self, container):
        try:
            self.results_ws.cell(row=1, column=100).value = ""
        except PermissionError:
            messagebox.showwarning("Unable to complete action", "The file you are trying to access is open in the "
                                                                "background.\nPlease make sure the file is closed "
                                                                "before performing this action")
        else:
            min_cell, max_cell = find_min_max_cell(self.factors_ws, 'A', str(400), 2000)
            self.best_results = [("", float("inf"), float("inf"))] * 5
            self.best_results = repeating_loop_best_factors(MIN_FACTOR, self.my_dict, self.factors_ws,
                                                            min_cell, max_cell, self.my_target, self.my_solar_spectrum,
                                                            self.number_of_loops.get(), self.best_results)
            count = 2
            for i in self.best_results:
                temp_dict = simulate_final_spectrum(self.my_dict, i, self.factors_ws, self.factor_places)
                put_results_in_sheet(self.results_ws, i, temp_dict, count)
                count += 1
                global_best_results.append(i)
            self.my_wb.save(self.label_file_path.get())

            if self.best_results_table is not None:
                self.best_results_table.destroy()
            self.best_results_table = tk.Frame(container)
            self.best_results_table.grid()
            self.table_label.grid()
            self.recommendation_chart.clear()
            self.recommendation_chart.append(tk.Label(
                self.best_results_table, text="Rank", font="Calibri 12", padx=2))
            self.recommendation_chart[-1].grid()
            self.recommendation_chart.append(tk.Label(
                self.best_results_table, text="Factors", font="Calibri 12", padx=2))
            self.recommendation_chart[-1].grid(row=0, column=1)
            self.recommendation_chart.append(tk.Label(
                self.best_results_table, text="Deviation", font="Calibri 12", padx=2))
            self.recommendation_chart[-1].grid(row=0, column=2)
            self.recommendation_chart.append(tk.Label(
                self.best_results_table, text="Alpha", font="Calibri 12", padx=2))
            self.recommendation_chart[-1].grid(row=0, column=3)
            self.recommendation_chart.append(tk.Label(
                self.best_results_table, text="Epsilon", font="Calibri 12", padx=2))
            self.recommendation_chart[-1].grid(row=0, column=4)
            self.recommendation_chart.append(tk.Label(
                self.best_results_table, text="Graph", font="Calibri 12", padx=2))
            self.recommendation_chart[-1].grid(row=0, column=5)
            for i, result in enumerate(self.best_results, 1):
                factors = factor_formatting(result[0])
                self.recommendation_chart.append(tk.Label(
                    self.best_results_table, text=i, font="Calibri 12", padx=10))
                self.recommendation_chart[-1].grid(row=i)
                self.recommendation_chart.append(tk.Label(
                    self.best_results_table, text=factors, font="Calibri 12", padx=2, anchor=tk.W))
                self.recommendation_chart[-1].grid(row=i, column=1, sticky="nsew")
                self.recommendation_chart.append(tk.Label(
                    self.best_results_table, text="%.4f" % result[1], font="Calibri 12", padx=2))
                self.recommendation_chart[-1].grid(row=i, column=2)
                self.recommendation_chart.append(tk.Label(
                    self.best_results_table, text=calculate_difference(result[2], self.original_alpha),
                    font="Calibri 12", padx=2))
                self.recommendation_chart[-1].grid(row=i, column=3)
                self.recommendation_chart.append(tk.Label(
                    self.best_results_table, text=calculate_difference(result[3], self.original_epsilon, 1),
                    font="Calibri 12", padx=2))
                self.recommendation_chart[-1].grid(row=i, column=4)
                self.recommendation_chart.append(tk.Button(
                    self.best_results_table, text="Show", padx=2, command=lambda r=result: self.open_graph_window(
                        simulate_final_spectrum(self.my_dict, r, self.factors_ws, self.factor_places),
                        factors, r[1], r[2], r[3])))
                self.recommendation_chart[-1].grid(row=i, column=5)

    # ----------------------------------------------------------------
    # Handles the "import" button click
    def import_button_clicked(self):
        self.root.filename = filedialog.askopenfilename(initialdir="/", title="Select file",
                                                        filetypes=("excel files", "*.xlsx"))

    # ----------------------------------------------------------------
    # Handles the factor buttons click
    def factor_button_clicked(self, factor):
        self.delete_factor_button.config(state="active")
        self.confirm_factor_button.config(state="active")
        new_text = self.text_factors.get(1.0, tk.END)[:-1]
        temp = new_text.split(", ")
        if temp[0] == '':
            new_text += factor
        elif len(temp) < MAX_NUM_OF_LOOPS:
            new_text += ", " + factor
        if len(temp) == MAX_NUM_OF_LOOPS - 1:
            for button in self.buttons:
                button.config(state="disabled")
        self.text_factors.delete(1.0, tk.END)
        self.text_factors.insert(1.0, new_text)

    # ----------------------------------------------------------------
    # Handles the "delete" button click
    def delete_factor_button_function(self):
        new_text = self.text_factors.get(1.0, tk.END)
        temp = new_text.split(", ")
        if len(temp) == 1:
            self.delete_factor_button.config(state="disabled")
            self.confirm_factor_button.config(state="disabled")
        temp.pop()
        new_text = ""
        for i, factor in enumerate(temp):
            new_text += factor
            if i < len(temp) - 1:
                new_text += ", "

        self.text_factors.delete(1.0, tk.END)
        self.text_factors.insert(1.0, new_text)

        for button in self.buttons:
            button.config(state="active")

    # ----------------------------------------------------------------
    # Handles the "confirm" button click
    def confirm_factor_button_function(self):
        factors = self.text_factors.get(1.0, tk.END)
        factors_list = factors.split(", ")
        factors_list[-1] = factors_list[-1][:-1]  # to remove the \n

        min_cell, max_cell = find_min_max_cell(self.factors_ws, 'A', str(400), 2000)

        ff = ""
        new_spectrum = self.my_dict
        for factor in factors_list:
            i = 1
            if "++" in factor:
                temp1 = column_index_from_string(self.factor_places[factor[:-2] + " +"])
                i = 2
                ff += factor[:-2] + " +" + ", " + factor[:-2] + " +" + ", "
            elif "+" in factor:
                temp1 = column_index_from_string(self.factor_places[factor[:-1] + " +"])
                ff += factor[:-1] + " +" + ", "
            elif "--" in factor:
                temp1 = column_index_from_string(self.factor_places[factor[:-2] + " -"])
                i = 2
                ff += factor[:-2] + " -" + ", " + factor[:-2] + " -" + ", "
            else:
                temp1 = column_index_from_string(self.factor_places[factor[:-1] + " -"])
                ff += factor[:-1] + " -" + ", "
            for _ in range(i):
                new_spectrum = simulate_spectrum(new_spectrum, self.factors_ws, temp1, min_cell, max_cell)
        ff = factor_formatting(ff[:-2], True)
        dev = compare_to_target(self.my_target, new_spectrum)
        alpha = calculate_alpha(self.my_solar_spectrum, new_spectrum)
        epsilon = calculate_epsilon(new_spectrum)
        self.open_graph_window(new_spectrum, ff, dev, alpha, epsilon)
        self.text_factors.delete('1.0', tk.END)
        self.delete_factor_button.config(state="disabled")
        self.confirm_factor_button.config(state="disabled")
        for button in self.buttons:
            button.config(state="active")

    # ----------------------------------------------------------------
    def optimize_target(self):
        min_cell, max_cell = find_min_max_cell(self.factors_ws, 'A', str(400), 2000)
        self.best_target_results = [("", float("inf"), float("inf"))] * 5
        self.best_target_results = repeating_loop_best_factors(MIN_FACTOR, self.my_dict, self.factors_ws, min_cell,
                                                               max_cell, self.my_target, self.my_solar_spectrum,
                                                               self.number_of_loops.get(), self.best_target_results,
                                                               evaluation_of_results=1)
        text = ""
        for num, i in enumerate(self.best_target_results):
            print("Factors used: ", factor_formatting(i[0]), "\nDeviation: %.4f" % i[1],
                  "\nAlpha change: %.2f" % i[2], "\nEpsilon change: %.1f" % i[3], "\n~~~~~~~~~~~~~~~~~~~\n")
            text += "Factors used: " + factor_formatting(i[0]) + "\nDeviation: %.4f" % i[1] + \
                    "\nAlpha change: %.2f" % i[2] + "\nEpsilon change: %.1f" % i[3] + "\n~~~~~~~~~~~~~~~~~~~\n"
        TextFrame(self, "Optimize target", text)

    def configure(self, _):
        center(self.root)

    # ----------------------------------------------------------------
    # Runs when the application closes
    def on_close(self):
        self.root.destroy()

        with open(self.new_file, 'wb') as fi:
            # dump your data into the file
            pickle.dump(self.label_file_path.get(), fi)


# ~~~~~~~~~~~~~~~~~~~~ Functions ~~~~~~~~~~~~~~~~~~~~~~

# \/\/\/ Helping Functions: \/\/\/

# centers the window on the screen
def center(top_level):
    top_level.update_idletasks()

    # Tkinter way to find the screen resolution
    screen_width = top_level.winfo_screenwidth()
    screen_height = top_level.winfo_screenheight()

    size = tuple(int(_) for _ in top_level.geometry().split('+')[0].split('x'))
    x = screen_width / 2 - size[0] / 2
    y = screen_height / 2 - size[1] / 2

    top_level.geometry("+%d+%d" % (x, y))


# Finds min_cell and max_cell of the range that we want to work with
def find_min_max_cell(worksheet, line='A', start="lambda [nm]", end=None):
    min_cell = 0
    max_cell = 0
    for cell in worksheet[line]:
        if str(cell.internal_value) == str(start):
            if start == "lambda [nm]":
                min_cell = cell.row + 1
            else:
                min_cell = cell.row
        if min_cell != 0:
            if end is None:
                if cell.internal_value is end:
                    max_cell = cell.row
                    break
            else:
                if str(cell.internal_value) == str(end):
                    max_cell = cell.row
                    break
    return min_cell, max_cell


def find_factors(factors_ws, factors_set, start="lambda [nm]"):
    factors_row = 0
    for cell in factors_ws['A']:
        if str(cell.internal_value) == start:
            factors_row = cell.row
            break
    for cell in factors_ws[factors_row]:
        if cell.internal_value is None:
            break
        elif str(cell.internal_value[-1]) == "+" or str(cell.internal_value[-1]) == "-":
            factors_set.add(str(cell.internal_value[:-2]))
    return factors_set


# Maps which column each factor resides in
def map_factor_places(factors_ws):
    # define variables
    temp_list = {}
    temp_row = 0
    for cell in factors_ws['A']:
        if str(cell.internal_value) == "lambda [nm]":
            temp_row = cell.row
            break
    # code
    for i in range(MIN_FACTOR, MAX_FACTOR):
        temp_list[factors_ws.cell(row=temp_row, column=i).internal_value] = \
            factors_ws.cell(row=temp_row, column=i).column
    return temp_list


# Organizes the Data into a dictionary
def data_reduce(worksheet, column=2, wavelength=400, max_wavelength=2000):
    # defining variables
    my_dict = {}
    closest_cell = 0.0
    min_cell, max_cell = find_min_max_cell(worksheet)
    # Fill my dictionary with relevant values from the Data sheet
    for row in range(min_cell, max_cell):
        current_cell = float(worksheet.cell(row=row, column=1).internal_value)
        if current_cell <= closest_cell:
            pass
        elif abs(wavelength - current_cell) <= abs(wavelength - closest_cell):
            closest_cell = current_cell
        elif wavelength < max_wavelength:
            my_dict[int(wavelength)] = worksheet.cell(row=row - 1, column=column).internal_value
            closest_cell = current_cell
            wavelength += 10
        if wavelength >= max_wavelength:
            break
    return my_dict


# Compares a given spectrum to the requested target, returns deviation
def compare_to_target(target, spectrum):
    # define variables
    dev = 0
    # code
    for i in target:
        dev += (float(target[i]) - spectrum[i]) ** 2
    return dev


def calculate_alpha(solar_spectrum, spectrum, showerror=False):
    alpha = 1
    sum1 = 0
    sum2 = 0
    for i in solar_spectrum:
        sum1 += float(solar_spectrum[i])
        sum2 += float(solar_spectrum[i]) * float(spectrum[i])
    try:
        alpha -= sum2 / sum1
    except ZeroDivisionError:
        if showerror:
            messagebox.showwarning("Alpha error", "could not calculate alpha properly.\nPlease check your solar_"
                                                  "spectrum sheet and data sheet")
    return alpha * 100


def calculate_epsilon(spectrum):
    lambda_min = 100000
    temp_ref = 1000
    for wl, ref in spectrum.items():
        if 1000 <= wl <= 1700:
            if temp_ref > ref:
                temp_ref = ref
                lambda_min = wl
    slope = (spectrum[1900] - spectrum[1700]) / (1900 - 1700) * 100
    result = 0.002 * lambda_min - 154.3 * slope + 17.5
    return result


def calculate_difference(variable, original, dec=2):
    if variable - original >= 0:
        return "+" + str(round(variable - original, dec))
    else:
        return str(round(variable - original, dec))


def factor_formatting(factors, is_spaced=False):
    if is_spaced:
        temp = 2
    else:
        temp = 1
    factors_dict = {}
    for i in factors.split(", "):
        factor = i[:-temp]
        if i[-temp:] == " +" or i[-temp:] == "+":
            sign = 1
        elif i[-temp:] == " -" or i[-temp:] == "-":
            sign = -1
        else:
            sign = 1000
        factors_dict.setdefault(factor, 0)
        factors_dict[factor] += sign
    formatted_text = ""
    is_empty = True
    for i in factors_dict:
        if factors_dict[i] != 0:
            if factors_dict[i] > 0:
                t = "+"
            else:
                t = ""
            if i[0] == "I":
                t += str(factors_dict[i]) + "A"
            elif i[0] == "V":
                t += ('%f' % float(factors_dict[i]/2)).rstrip('0').rstrip('.') + "mm/s"
            else:
                t += str(factors_dict[i])
            formatted_text += i + ": " + t + ",  "
            is_empty = False
    if is_empty:
        formatted_text = "no change"
    return formatted_text


# find spectrum from factor names
# def spectrum_from_factors(factors):

# \/\/\/ Main Functions: \/\/\/

# Simulates the expected spectrum that will be achieved by using the function
# in the given column on the data in the dict
def simulate_spectrum(my_dict, factors_ws, column, min_cell, max_cell):
    # defining variables
    temp_dict = {}
    # code
    for i in range(int(min_cell), int(max_cell)):
        temp = factors_ws.cell(row=i, column=column_index_from_string('A')).internal_value
        temp_dict[temp] = my_dict[temp] + (factors_ws.cell(row=i, column=column).internal_value / 100.0)
    return temp_dict


# Uses simulate_spectrum a number of times with the requested factors in the result
def simulate_final_spectrum(my_dict, result, factors_ws, factor_to_column):
    words = result[0].split(", ")
    temp_dict = my_dict.copy()
    min_cell, max_cell = find_min_max_cell(factors_ws, 'A', str(400), 2000)
    for i in range(0, len(words)):
        temp_dict = simulate_spectrum(temp_dict, factors_ws, int(column_index_from_string(factor_to_column[words[i]])),
                                      min_cell, max_cell)
    return temp_dict


# The main function, recursively repeats a requested number of loops
# that run on the factor sheet and simulate all the possible spectrums
def repeating_loop_best_factors(min_factor, spectrum, factors_ws, min_cell, max_cell, my_target, solar_spectrum,
                                num_of_loops, best_results, loops_left=-1, used_factors=None, evaluation_of_results=0):
    # on first iteration only
    if loops_left == -1:
        used_factors = [""] * num_of_loops
        loops_left = num_of_loops
    # checks when to stop
    if loops_left > 0:
        # main loop
        for i in range(min_factor, MAX_FACTOR):
            # keeps array of iterators to know how to append relevant factor names
            used_factors[num_of_loops - loops_left] = factors_ws.cell(row=5, column=i).internal_value
            # creates the new spectrum
            new_spectrum = simulate_spectrum(spectrum, factors_ws, i, min_cell, max_cell)
            # calls this function again with the new values and one less loops left
            best_results = repeating_loop_best_factors(i, new_spectrum, factors_ws, min_cell, max_cell, my_target,
                                                       solar_spectrum, num_of_loops, best_results, loops_left - 1,
                                                       used_factors)
            # on the last active loop
            if loops_left == 1:
                text = ""
                for j in used_factors:
                    text += j + ", "
                text = text[0:-2]
                # checks deviation score and checks if it is in the top results
                dev = compare_to_target(my_target, new_spectrum)
                alpha = calculate_alpha(solar_spectrum, new_spectrum)
                epsilon = calculate_epsilon(spectrum)
                best_results.append((text, dev, alpha, epsilon))
                if evaluation_of_results == 0:
                    best_results = sorted(best_results, key=itemgetter(1), reverse=False)[:5]
                else:
                    print("hello!")
    return best_results


# Puts the best results into the result sheet
def put_results_in_sheet(results_ws, result, my_dict, count):
    # define variables
    min_cell, max_cell = find_min_max_cell(results_ws, 'A', str(400), 2000)
    factors = None
    for cell in results_ws['A']:
        if str(cell.internal_value) == "factors":
            factors = cell.row
            break
    # code
    results_ws.cell(row=int(factors), column=count).value = result[0]
    results_ws.cell(row=int(factors + 1), column=count).value = round(Decimal(result[1]), 4)
    for i in range(int(min_cell), int(max_cell)):
        results_ws.cell(row=i, column=count).value = round(Decimal(my_dict[results_ws.cell(row=i, column=1).value]), 8)


if __name__ == "__main__":
    root = tk.Tk()
    app = AlgosApp(root)
    root.mainloop()
